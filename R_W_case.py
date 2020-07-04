from openpyxl import Workbook
from openpyxl import load_workbook
def read_data(file_name,sheet_name):
    wb=load_workbook(file_name)  #打开一个工作簿
    sheet=wb[sheet_name]     #工作簿变量名【表单名】
    data = []
    for i in range(2,sheet.max_row+1):      #行增加
        a=[]
        for j in range(1,sheet.max_column-1):                            #列增加
            value = sheet.cell(row=i,column=j).value
            a.append(value)
        data.append(a)
    return data
def write_data(file_name,sheet_name,row,column,value):
    wb = load_workbook(file_name)
    sheet = wb[sheet_name]
    sheet.cell(row= row, column=column).value = value
    wb.save('Mycase.xlsx')

if __name__ == '__main__':                     #在当前页面执行才可以
    all_data=read_data('Mycase.xlsx','recharge')
    print("所有的测试数据：",all_data)

